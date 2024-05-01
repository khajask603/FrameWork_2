import openpyxl


class HomePagedata:

    test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]


    @staticmethod
    def getTestdata(testcasename):
        Dict = {}
        filePath = r"F:\PHYTON&SeleniumByRS\Frameworks\PhytonSelFramework\pythonProject\TestData\PhytonDemo.xlsx"
        workbook = openpyxl.load_workbook(filePath)
        # sheet=workbook.active
        sheet = workbook["Sheet1"]  # -->Diffrent approach on pulling the sheets

        rows = sheet.max_column
        colms = sheet.max_column
        testcase_row = 0
        for r in range(1, rows + 1):
            if sheet.cell(r, column=1).value == testcasename:
                testcase_row = r  # Storing the row number where "TestcaseName" is found
                break

        for c in range(testcase_row, colms + 1):
            # Dict=["FirstName"]="Rahul"                                ->Expeceted way we want to pull the data on below code line
            # Dict[sheet.cell(row=1, column=c).value] = sheet.cell(testcase_row, c).value
            key = sheet.cell(row=1, column=c).value.strip()  # Removing leading and trailing spaces from the key
            value = sheet.cell(testcase_row, c).value
            Dict[key]=value
        return[Dict]