import openpyxl

filePath = r"F:\PHYTON&SeleniumByRS\Frameworks\PhytonSelFramework\pythonProject\TestData\PhytonDemo.xlsx"
workbook = openpyxl.load_workbook(filePath)
# sheet=workbook.active
sheet = workbook["Sheet1"]  # -->Diffrent approach on pulling the sheets
Dict = {}

# 1)----Printing specific Value from specific row & Column
cell = sheet.cell(row=1, column=2)
print(cell.value)

# 2)----Writing & Printing specific Value from specific row & Column

sheet.cell(row=2, column=2).value = "Rahul"
print(sheet.cell(row=2, column=2).value)

# 3)--Printing count of Avialble total number of rows and columns
print(sheet.max_row)
print(sheet.max_column)

# 4--) Printing all the data from excel sheet

rows = sheet.max_column
colms = sheet.max_column

print("---------------Values from row with Testcase2------------------------")
testcase_row=0
for r in range(1, rows + 1):
    if sheet.cell(r, column=1).value == "Testcase2":
        testcase_row = r  # Store the row number where "Testcase2" is found
        break

for c in range(testcase_row, colms + 1):
    # Dict=["FirstName"]="Rahul"                                ->Expeceted way we want to pull the data on below code line
    Dict[sheet.cell(row=1, column=c).value] = sheet.cell(testcase_row, c).value
    print(sheet.cell(testcase_row, c).value, end='   ||     ')
    # print(sheet.cell(r,c).value)
print()
print(Dict)
# -->COnclusion Comands EveryTime-.Savinf--
# workbook.save(filePath)
